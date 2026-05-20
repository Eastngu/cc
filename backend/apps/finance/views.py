from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from apps.users.permissions import IsBossOrFinance
from .models import Receivable, Payable, Payment, MonthlyStatement
from .serializers import (
    ReceivableSerializer, PayableSerializer, PaymentSerializer,
    MonthlyStatementListSerializer, MonthlyStatementDetailSerializer,
)


class ReceivableViewSet(viewsets.ModelViewSet):
    queryset = Receivable.objects.select_related('customer').all()
    serializer_class = ReceivableSerializer
    permission_classes = [IsBossOrFinance]
    filterset_fields = ['customer', 'status', 'year', 'month']
    search_fields = ['receivable_no', 'customer__name', 'customer__short_name']
    ordering_fields = ['total_amount', 'balance', 'due_date', 'created_at']


class PayableViewSet(viewsets.ModelViewSet):
    queryset = Payable.objects.all()
    serializer_class = PayableSerializer
    permission_classes = [IsBossOrFinance]
    filterset_fields = ['status', 'category']
    search_fields = ['payable_no', 'supplier_name']
    ordering_fields = ['total_amount', 'balance', 'due_date']


class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.select_related('customer', 'receivable', 'payable').all()
    serializer_class = PaymentSerializer
    permission_classes = [IsBossOrFinance]
    filterset_fields = ['type', 'customer', 'payment_method']
    search_fields = ['payment_no']
    ordering_fields = ['amount', 'payment_date', 'created_at']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class MonthlyStatementViewSet(viewsets.ModelViewSet):
    queryset = MonthlyStatement.objects.select_related('customer', 'confirmed_by').all()
    permission_classes = [IsBossOrFinance]
    filterset_fields = ['customer', 'status', 'year', 'month']
    search_fields = ['statement_no', 'customer__name', 'customer__short_name']

    def get_serializer_class(self):
        if self.action == 'list':
            return MonthlyStatementListSerializer
        return MonthlyStatementDetailSerializer

    @action(detail=False, methods=['post'], url_path='generate')
    def generate(self, request):
        """Generate statement for a customer+month by pulling all shipped orders."""
        customer_id = request.data.get('customer')
        year = request.data.get('year')
        month = request.data.get('month')

        if not all([customer_id, year, month]):
            return Response({'detail': '请提供客户、年份和月份'}, status=400)

        year, month = int(year), int(month)

        if MonthlyStatement.objects.filter(customer_id=customer_id, year=year, month=month).exists():
            return Response({'detail': '该客户本月对账单已存在'}, status=400)

        from apps.orders.models import Order
        orders = Order.objects.filter(
            customer_id=customer_id,
            status='shipped',
            shipped_at__year=year,
            shipped_at__month=month,
        )

        total = orders.aggregate(total=Sum('total_amount'))['total'] or 0

        statement = MonthlyStatement.objects.create(
            customer_id=customer_id,
            year=year,
            month=month,
            total_amount=total,
        )
        statement.orders.set(orders)

        serializer = MonthlyStatementDetailSerializer(statement)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=['get'], url_path='export')
    def export_excel(self, request, pk=None):
        """Export statement as an Excel file."""
        statement = self.get_object()
        orders = statement.orders.all().order_by('shipped_at')

        wb = Workbook()
        ws = wb.active
        ws.title = '对账单'

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12

        # Styles
        title_font = Font(size=16, bold=True)
        header_font = Font(size=10, bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Row 1: Title
        ws.merge_cells('A1:H1')
        ws['A1'] = '月度对账单'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Rows 3-5: Statement info
        ws['A3'] = f'客户: {statement.customer.name}'
        ws['A4'] = f'账期: {statement.year}年{statement.month}月'
        ws['A5'] = f'对账单号: {statement.statement_no}'
        ws['E3'] = f'状态: {statement.get_status_display()}'

        # Row 7: Table header
        headers = ['订单号', '产品名称', '规格型号', '数量', '单位', '单价', '金额', '出货日期']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Rows 8+: Order data
        row = 8
        for order in orders:
            ws.cell(row=row, column=1, value=order.order_no).border = thin_border
            ws.cell(row=row, column=2, value=order.product_name).border = thin_border
            ws.cell(row=row, column=3, value=order.product_spec).border = thin_border
            ws.cell(row=row, column=4, value=float(order.quantity)).border = thin_border
            ws.cell(row=row, column=5, value=order.unit).border = thin_border
            ws.cell(row=row, column=6, value=float(order.unit_price)).border = thin_border
            ws.cell(row=row, column=7, value=float(order.total_amount)).border = thin_border
            ws.cell(
                row=row, column=8,
                value=str(order.shipped_at) if order.shipped_at else ''
            ).border = thin_border
            row += 1

        # Footer: totals (leave one blank row after data)
        row += 1
        ws.cell(row=row, column=6, value='合计:').font = header_font
        ws.cell(row=row, column=7, value=float(statement.total_amount)).font = header_font
        row += 1
        ws.cell(row=row, column=6, value='调整:').font = header_font
        ws.cell(row=row, column=7, value=float(statement.adjustment))
        row += 1
        ws.cell(row=row, column=6, value='最终金额:').font = header_font
        ws.cell(row=row, column=7, value=float(statement.final_amount)).font = Font(size=12, bold=True)

        # Build HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'对账单_{statement.statement_no}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=True, methods=['patch'], url_path='confirm')
    def confirm(self, request, pk=None):
        """Confirm a statement and auto-create a receivable."""
        statement = self.get_object()
        if statement.status != 'draft':
            return Response({'detail': '只有草稿状态可以确认'}, status=400)

        statement.status = 'confirmed'
        statement.confirmed_by = request.user
        statement.confirmed_at = timezone.now()
        statement.save()

        customer = statement.customer
        import datetime
        due_date = timezone.now().date() + datetime.timedelta(days=customer.payment_terms)

        receivable, created = Receivable.objects.get_or_create(
            customer=customer,
            year=statement.year,
            month=statement.month,
            defaults={
                'total_amount': statement.final_amount,
                'due_date': due_date,
            }
        )
        if not created:
            receivable.total_amount = statement.final_amount
            receivable.save()

        serializer = MonthlyStatementDetailSerializer(statement)
        return Response(serializer.data)
